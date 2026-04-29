from io import StringIO
from pathlib import Path
import csv

from textual import on
from textual.app import App, ComposeResult
from textual.containers import Horizontal, Vertical, VerticalScroll
from textual.reactive import reactive
from textual.widgets import Button, Footer, Header, Label, Select, Static

from plot_widget import DataPlot
from tree import FilteredDirectoryTree

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class EdaExplorerApp(App):
    """EDA from your terminal, with file loading and plotting."""

    CSS_PATH = "style.tcss"

    BINDINGS = [
        ("q", "quit", "Quit"),
    ]

    selected_file = reactive(None)
    active_csv_text = reactive("")
    column_names = reactive([])
    loaded_rows = reactive([])

    def compose(self) -> ComposeResult:
        yield Header()
        with Horizontal(id="main"):
            yield FilteredDirectoryTree("./data", id="file-tree")
            with Vertical(id="right-panel"):
                yield Label(
                    "Select a .csv or .xlsx file from the tree on the left",
                    id="status",
                )
                with Horizontal(id="button-row"):
                    yield Button("Load / Convert", id="load-btn", variant="primary")
                    yield Button("Overview", id="stats-overview-btn")
                    yield Button("Missing", id="stats-missing-btn")
                    yield Button("Headers", id="stats-headers-btn")
                with Horizontal(id="lower-panel"):
                    with Vertical(id="controls-panel"):
                        yield Label("X Axis", classes="axis-label")
                        yield Select(options=[], id="x-col")
                        yield Label("Y Axis", classes="axis-label")
                        yield Select(options=[], id="y-col")
                        yield Label("Plot Type", classes="axis-label")
                        yield Select(
                            options=[
                                ("Line", "line"),
                                ("Bar", "bar"),
                                ("Scatter", "scatter"),
                                ("Histogram", "hist"),
                            ],
                            value="line",
                            id="plot-type",
                        )
                        yield Button(
                            "Plot",
                            id="plot-btn",
                            variant="success",
                            disabled=True,
                        )
                    with Vertical(id="display-panel"):
                        yield Label("File Stats", classes="axis-label")
                        with VerticalScroll(id="stats-panel"):
                            yield Static(
                                "Load a file to inspect summary statistics.",
                                id="stats-output",
                            )
                        yield DataPlot(id="data-plot")
        yield Footer()

    def watch_selected_file(self, path: Path | None) -> None:
        self.active_csv_text = ""
        self.column_names = []
        self.loaded_rows = []
        self.query_one("#plot-btn", Button).disabled = True
        self.query_one("#stats-output", Static).update(
            "Load a file to inspect summary statistics."
        )

        if path:
            self.query_one("#status", Label).update(
                f"{path.name} - press 'Load / Convert' to inspect"
            )

    def watch_column_names(self, names: list) -> None:
        options = [(name, name) for name in names]
        x_select = self.query_one("#x-col", Select)
        y_select = self.query_one("#y-col", Select)
        x_select.set_options(options)
        y_select.set_options(options)

        if names:
            x_default, y_default = self._default_axis_columns(names)
            x_select.value = x_default
            y_select.value = y_default

        self.query_one("#plot-btn", Button).disabled = len(names) == 0

    def _load_csv_text(self, path: Path) -> str:
        return path.read_text(encoding="utf-8", errors="replace")

    def _load_xlsx_text(self, path: Path) -> str:
        if pd is not None:
            dataframe = pd.read_excel(path)
            return dataframe.to_csv(index=False)

        if load_workbook is None:
            raise RuntimeError(
                "Loading .xlsx files requires pandas or openpyxl to be installed"
            )

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        output = StringIO()
        writer = csv.writer(output, lineterminator="\n")

        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])

        workbook.close()
        return output.getvalue()

    def _parse_rows(self, csv_text: str) -> list[dict[str, str]]:
        reader = csv.DictReader(csv_text.splitlines())
        rows = []
        for row in reader:
            cleaned_row = {}
            for key, value in row.items():
                if key is None:
                    continue
                cleaned_row[key] = "" if value is None else str(value).strip()
            rows.append(cleaned_row)
        return rows

    def _numeric_values(self, column_name: str) -> list[float]:
        numeric = []
        for row in self.loaded_rows:
            value = row.get(column_name, "")
            if value == "":
                continue
            try:
                numeric.append(self._parse_numeric(value))
            except ValueError:
                continue
        return numeric

    def _parse_numeric(self, value: str) -> float:
        cleaned = value.strip().replace(",", "")
        if cleaned == "":
            raise ValueError("empty numeric value")
        return float(cleaned)

    def _is_numeric_column(self, column_name: str) -> bool:
        values = [row.get(column_name, "") for row in self.loaded_rows if row.get(column_name, "") != ""]
        if not values:
            return False
        return len(self._numeric_values(column_name)) == len(values)

    def _show_stats_message(self, message: str) -> None:
        self.query_one("#stats-output", Static).update(message)

    def _build_dataframe(self):
        if pd is None or not self.loaded_rows:
            return None

        return pd.DataFrame(self.loaded_rows)

    def _build_overview_text(self) -> str:
        lines = [
            f"Rows: {len(self.loaded_rows)}",
            f"Columns: {len(self.column_names)}",
            "",
            "Column summary:",
        ]

        for name in self.column_names:
            values = [row.get(name, "") for row in self.loaded_rows]
            missing = sum(1 for value in values if value == "")
            non_missing = len(values) - missing
            numeric = self._numeric_values(name)

            if numeric:
                mean = sum(numeric) / len(numeric)
                lines.append(
                    f"{name}: non-missing={non_missing}, missing={missing}, "
                    f"numeric={len(numeric)}, min={min(numeric):.2f}, "
                    f"max={max(numeric):.2f}, mean={mean:.2f}"
                )
            else:
                unique_count = len({value for value in values if value != ""})
                lines.append(
                    f"{name}: non-missing={non_missing}, missing={missing}, "
                    f"unique={unique_count}"
                )

        return "\n".join(lines)

    def _infer_column_type(self, column_name: str) -> str:
        values = [row.get(column_name, "") for row in self.loaded_rows if row.get(column_name, "") != ""]
        if not values:
            return "empty"

        numeric = self._numeric_values(column_name)
        if len(numeric) == len(values):
            return "numeric"

        unique_count = len(set(values))
        if unique_count <= max(20, len(values) // 10):
            return "categorical"

        if numeric:
            return "mixed"

        return "text"

    def _graph_usage_for_type(self, column_type: str) -> str:
        if column_type == "numeric":
            return "Y: line/bar/scatter, values: histogram, X: line/scatter"
        if column_type == "categorical":
            return "X: bar/line/scatter"
        if column_type == "mixed":
            return "X: bar/line/scatter, Y: maybe after cleaning"
        if column_type == "text":
            return "X: bar/line/scatter if used as labels"
        return "Needs cleaning before plotting"

    def _build_header_text(self) -> str:
        lines = [f"All headers ({len(self.column_names)}):", ""]

        for index, name in enumerate(self.column_names, start=1):
            column_type = self._infer_column_type(name)
            graph_usage = self._graph_usage_for_type(column_type)
            lines.append(
                f"{index}. {name}\n"
                f"   type: {column_type}\n"
                f"   graphs: {graph_usage}"
            )

        return "\n".join(lines)

    def _default_axis_columns(self, names: list[str]) -> tuple[str, str]:
        x_default = names[0]

        numeric_names = [name for name in names if self._numeric_values(name)]
        if numeric_names:
            y_default = numeric_names[0]
            if y_default == x_default and len(numeric_names) > 1:
                y_default = numeric_names[1]
        else:
            y_default = names[1] if len(names) > 1 else names[0]

        if x_default == y_default and len(names) > 1:
            for name in names:
                if name != x_default:
                    y_default = name
                    break

        return x_default, y_default

    @on(Button.Pressed, "#load-btn")
    def load_file(self) -> None:
        if self.selected_file is None:
            self.query_one("#status", Label).update("Select a file first")
            return

        path = Path(self.selected_file)
        self.query_one("#status", Label).update(f"Loading {path.name} ...")

        try:
            if path.suffix.lower() == ".xlsx":
                csv_text = self._load_xlsx_text(path)
            elif path.suffix.lower() == ".csv":
                csv_text = self._load_csv_text(path)
            else:
                self.query_one("#status", Label).update(
                    f"Unsupported file type: {path.suffix}"
                )
                return

            reader = csv.DictReader(csv_text.splitlines())
            names = reader.fieldnames or []

            if not names:
                self.active_csv_text = ""
                self.column_names = []
                self.query_one("#status", Label).update(
                    f"{path.name} could not be read as tabular data"
                )
                return

            self.active_csv_text = csv_text
            self.column_names = names
            self.loaded_rows = self._parse_rows(csv_text)
            row_count = len(self.loaded_rows)

            self.query_one("#status", Label).update(
                f"{path.name} | {len(names)} columns | ~{row_count} rows"
                " - pick X and Y columns, then press 'Plot'"
            )
            self._show_stats_message(
                "Press Overview, Missing, or Headers to inspect this file."
            )
        except Exception as exc:
            self.query_one("#status", Label).update(f"Error: {exc}")

    @on(Button.Pressed, "#stats-overview-btn")
    def show_overview_stats(self) -> None:
        if not self.loaded_rows:
            self._show_stats_message("Load a file first.")
            return

        self._show_stats_message(self._build_overview_text())

    @on(Button.Pressed, "#stats-missing-btn")
    def show_missing_stats(self) -> None:
        if not self.loaded_rows:
            self._show_stats_message("Load a file first.")
            return

        parts = []
        for name in self.column_names[:8]:
            missing = sum(1 for row in self.loaded_rows if row.get(name, "") == "")
            parts.append(f"{name}: {missing}")

        extra = ""
        if len(self.column_names) > 8:
            extra = "\nMore columns omitted."

        self._show_stats_message(
            "Missing values by column:\n" + "\n".join(parts) + extra
        )

    @on(Button.Pressed, "#stats-headers-btn")
    def show_header_stats(self) -> None:
        if not self.loaded_rows:
            self._show_stats_message("Load a file first.")
            return

        self._show_stats_message(self._build_header_text())

    @on(Button.Pressed, "#plot-btn")
    def plot_selected(self) -> None:
        x_col = self.query_one("#x-col", Select).value
        y_col = self.query_one("#y-col", Select).value
        plot_type = str(self.query_one("#plot-type", Select).value)

        if y_col is Select.BLANK:
            self.query_one("#status", Label).update(
                "Choose a Y column before plotting"
            )
            return

        if plot_type != "hist" and x_col is Select.BLANK:
            self.query_one("#status", Label).update(
                "Choose both an X and a Y column before plotting"
            )
            return

        if not self.active_csv_text:
            self.query_one("#status", Label).update("Load a file first")
            return

        y_col = str(y_col)
        x_col = str(x_col)
        reader = csv.DictReader(self.active_csv_text.splitlines())
        x_values = []
        y_values = []
        skipped = 0

        for row in reader:
            raw_y = row.get(y_col, "").strip()
            raw_x = row.get(x_col, "").strip()
            try:
                y_values.append(self._parse_numeric(raw_y))
                try:
                    x_values.append(float(raw_x))
                except ValueError:
                    x_values.append(raw_x)
            except ValueError:
                skipped += 1

        self.query_one(DataPlot).replot_xy(
            x_col, y_col, x_values, y_values, plot_type
        )

        note = f" ({skipped} row(s) skipped - non-numeric Y)" if skipped else ""
        self.query_one("#status", Label).update(
            f"Plotted: {y_col} vs {x_col}{note}"
        )


if __name__ == "__main__":
    EdaExplorerApp().run()
